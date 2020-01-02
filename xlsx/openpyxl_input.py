# -*- coding:utf-8 -*-
from openpyxl import Workbook

'''
写入excel
'''


def write_excel():
    wb = Workbook()  # 打开 excel 工作簿
    ws1 = wb.active  # 激活
    # ws1.title = ""  # 给予文件标题
    # 存第一行单元格cell(1,1)
    ws1.cell(1, 1).value = 6  # 这个方法索引从1开始
    ws1.cell(2, 1).value = 6  # 这个方法索引从1开始
    ws1.cell(3, 1).value = 6  # 这个方法索引从1开始
    ws1['A4'] = 4

    # 存一行数据
    ws1.append([11, 87])
    wb.save("openpyxl.xlsx")


if __name__ == '__main__':
    write_excel()
    print("openyxl已将数据写入Excel")

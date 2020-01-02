# -*- coding:utf-8 -*-
from openpyxl import load_workbook

def read_excel():
    import openpyxl

    wb = openpyxl.load_workbook('openpyxl.xlsx')
    # 获取所有工作表名
    names = wb.sheetnames
    # wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
    sheet = wb[names[0]]
    # 获取最大行数
    maxRow = sheet.max_row
    # 获取最大列数
    maxColumn = sheet.max_column
    # 获取当前活动表
    current_sheet = wb.active
    # 获取当前活动表名称
    current_name = sheet.title
    # 通过名字访问Cell对象, 通过value属性获取值
    a1 = sheet['A1'].value
    print("a1的数据:{}".format(a1))
    # 通过行和列确定数据
    a52 = sheet.cell(row=5, column=2).value
    print("a5 2的数据:{}".format(a52))
    # 获取列字母
    column_name = openpyxl.utils.cell.get_column_letter(1)
    # 将列字母转为数字, 参数忽略大小写
    column_name_num = openpyxl.utils.cell.column_index_from_string('a')
    # 获取一列数据, sheet.iter_rows() 获取所有的行
    for one_column_data in sheet.iter_rows():
        print(one_column_data[0].value)
    # for one_row_data in sheet.iter_cols():
    #     print(one_row_data[0].value, end="\t")

    print("row = {}, column = {}".format(maxRow, maxColumn))



    # workbook = load_workbook('openpyxl.xlsx')
    # # booksheet = workbook.active                #获取当前活跃的sheet,默认是第一个sheet
    # sheets = workbook.get_sheet_names()  # 从名称获取sheet
    # booksheet = workbook.get_sheet_by_name(sheets[0])
    #
    # rows = booksheet.rows
    # columns = booksheet.columns
    # # 迭代所有的行
    # for row in rows:
    #     line = [col.value for col in row]
    #
    # # 通过坐标读取值
    # cell_11 = booksheet.cell('A1')
    # cell_11 = booksheet.cell(row=1, column=1).value

if __name__ == '__main__':
    read_excel()
    print("openpyxl已将Excel读取完毕")